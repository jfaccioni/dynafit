from operator import attrgetter
from typing import Tuple

import scipy.interpolate as interpolate
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

SETTINGS = {
    # path to input file.
    'path': 'data/Pasta para Ju.xlsx',
    # Colony Size cell range in input Excel file
    'cs_range': 'A4:A1071',
    # Growth Rate range in input Excel file
    'gr_range': 'B4:B1071',
    # max size of colony to not be binned, e.g. a parameter of 10 makes colonies of up to 10 cells to be
    # plotted individually in CVP, instead of being binned along with other cell sizes.
    'max_binned_colony_size': 10,
    # number of bins in which to divide the population after the max number of individual colony sizes
    # determined by the parameter before.
    'bins': 5,
    # number of independent runs.
    'runs': 10,
    # number of repeated samplings to perform for each run.
    'repeats': 10,
    # number of instances in each sample.
    'sample_size': 20,
    # whether to interpolate mean value in CVP.
    'smoothing': False,
    # whether to show a histogram as well
    'show_hist': True,
}


def main(path: str, cs_range: str, gr_range: str, max_binned_colony_size: int, bins: int, runs: int, repeats: int,
         sample_size: int, smoothing: bool, show_hist: True) -> None:
    """Main function of this script"""
    fig, ax = plt.subplots(figsize=(16, 16))
    data = load_data(path=path, cs_range=cs_range, gr_range=gr_range)
    data = add_bins(data=data, max_binned_colony_size=max_binned_colony_size, bins=bins)
    for _ in range(runs):
        cs, gr = sample_data(data=data, repeats=repeats, sample_size=sample_size)
        ax.plot(cs, gr, color='black', alpha=0.5, marker='.')
    title = get_plot_title(runs=runs, repeats=repeats, sample_size=sample_size)
    format_plot(fig, ax, title, smoothing)
    if show_hist is True:
        plot_histogram(data=data)


def load_data(path: str, cs_range: str, gr_range: str) -> pd.DataFrame:
    """Loads relevant data from input file as a Pandas DataFrame.
    ASSUMES: data to be in cell range A4 to B1071; modify this as needed"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    values = [(cs[0].value, gr[0].value) for cs, gr in zip(ws[cs_range], ws[gr_range])]
    return pd.DataFrame(values, columns=('CS1', 'GR2'))


def add_bins(data: pd.DataFrame, max_binned_colony_size: int, bins: int) -> pd.DataFrame:
    """Returns the data DataFrame with a "bins" column, which divides the population of values in
    bins with a close number of instances in them"""
    data['bins'] = data['CS1']
    bin_condition = data['bins'] > max_binned_colony_size
    binned_data = pd.qcut(data.loc[bin_condition]['CS1'], bins, labels=False)
    binned_data += max_binned_colony_size + 1
    binned_data = pd.concat([binned_data, data.loc[~bin_condition]['bins']])
    return data.assign(bins=binned_data)


def sample_data(data: pd.DataFrame, repeats: int, sample_size: int) -> Tuple[np.ndarray, np.ndarray]:
    """Samples data and returns a tuple of arrays ready for plotting. This accounts for any
    necessary transformations before plotting, like log2."""
    sampling_result = random_sampling(data, repeats, sample_size)
    cs = np.log2(sampling_result.index)
    gr = sampling_result.mean(axis=1)  # means of each repeat across columns
    return cs, gr


def random_sampling(data: pd.DataFrame, repeats: int, sample_size: int) -> pd.DataFrame:
    """Samples each bin in the population. This is performed for a specific number of times (repeats),
    and each sampling contains a specific number of instances (sample_size)."""
    index = data.groupby('bins').mean()['CS1']
    output = pd.DataFrame(index=index)
    for i in range(repeats):
        variances = []
        for groupname, group in data.groupby('bins'):
            sample = group.sample(n=sample_size)
            variances.append(sample['GR2'].var())
        output[f'repeat_{i+1}'] = variances
    return output.applymap(np.log2)


def get_plot_title(runs, repeats, sample_size) -> str:
    """Returns a string for the plot's title."""
    return (f'CVP\n'
            f'independent runs: {runs}\n'
            f'sampling repeats: {repeats}\n'
            f'sample size: {sample_size}')


def format_plot(fig: plt.Figure, ax: plt.Axes, title: str, smoothing: bool) -> None:
    """Adds formatting to CVP."""
    fig.suptitle(title)
    ax.set_xlabel('log2(colony size)')
    ax.set_ylabel('log2(variance)')
    set_limits(ax)
    plot_mean_line(ax, smoothing)
    plot_supporting_lines(ax)


def set_limits(ax: plt.Axes) -> None:
    """Sets ax limits based on max and min values of XY lines."""
    _, max_x, min_x = get_axes_params(ax, 'x')
    _, max_y, min_y = get_axes_params(ax, 'y')
    ax.set_xlim(min_x - (abs(min_x) * 0.05), max_x + (abs(max_x) * 0.05))
    ax.set_ylim(min_y - (abs(min_y) * 0.05), max_y + (abs(max_y) * 0.05))


def plot_supporting_lines(ax: plt.Axes) -> None:
    """Plots the three supporting lines of the CVP."""
    start_x, *_ = get_axes_params(ax, 'x')
    start_y, *_ = get_axes_params(ax, 'y')
    ax.axhline(start_y, color='blue', lw=3)
    ax.plot([start_x, 10], [start_y, start_y - 10], color='red', lw=3)
    ax.axvline(start_x, color='black', lw=3)


def plot_mean_line(ax: plt.Axes, smoothing: bool) -> None:
    """Plots mean line for all data in ax."""
    xs = [x for x in ax.lines[0].get_xdata()]
    ys = np.array([line.get_ydata() for line in ax.lines]).mean(axis=0)
    ax.plot(xs, ys, color='green', alpha=0.9, lw=3)
    if smoothing is True:
        smooth_xs, smooth_ys = perform_smoothing(np.array(xs), ys)
        ax.plot(smooth_xs, smooth_ys, color='purple', alpha=0.9, lw=3)


def perform_smoothing(xs: np.ndarray, ys: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    """Performs smoothing with scipy.interpolate"""
    # noinspection PyArgumentList
    new_xs = np.linspace(xs.min(), xs.max(), 100)
    spline = interpolate.BSpline(*interpolate.splrep(xs, ys, s=5, k=4), extrapolate=False)
    return new_xs, spline(new_xs)


def get_axes_params(ax: plt.Axes, coord: str) -> Tuple[float, float, float]:
    """Gets start, max and min position of all lines in ax for X or Y coordinate."""
    coord_data = attrgetter(f'get_{coord}data')
    start_coord = max([coord_data(line)()[0] for line in ax.lines])
    max_coord = max([v for line in ax.lines for v in coord_data(line)()])
    min_coord = min([v for line in ax.lines for v in coord_data(line)()])
    return start_coord, max_coord, min_coord


def plot_histogram(data: pd.DataFrame) -> None:
    """Plots a histogram of the colony size, indicating the "cuts" made by the binning process"""
    fig, ax = plt.subplots()
    grouped_data = data.groupby('bins')
    ax.hist(np.log2(data['CS1']))
    ax.set_xlabel('log2(colony size)')
    ax.set_ylabel('count')
    ax.set_title('Colony size histogram')
    for xmax, label in zip(grouped_data.max()['CS1'], grouped_data.count()['CS1']):
        ax.axvline(np.log2(xmax), c='k')
        ax.text(np.log2(xmax), ax.get_ylim()[1] * 0.9, f'{xmax}\nn={label}')
    plt.show()


if __name__ == '__main__':
    main(**SETTINGS)
