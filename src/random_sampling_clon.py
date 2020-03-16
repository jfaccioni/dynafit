from src import random_sampling

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd

SETTINGS = {
    'path': '../data/clonogenico CS1 GR2.xlsx',
    'start_on_row': 3,
    'max_binned_colony_size': 0,
    'bins': 6,
    'runs': 10,
    'repeats': 10,
    'sample_size': 20,
    'smoothing': True,
}


def main(path: str, start_on_row: int, max_binned_colony_size: int, bins: int, runs: int, repeats: int,
         sample_size: int, smoothing: bool) -> None:
    """Main function of this script"""
    wb = openpyxl.load_workbook(path)
    for ws_name in wb.sheetnames:
        try:
            fig, ax = plt.subplots(figsize=(16, 16))
            data = load_clean_data(ws=wb[ws_name], n=start_on_row)
            data = random_sampling.add_bins(data=data, max_binned_colony_size=max_binned_colony_size, bins=bins)
            for _ in range(runs):
                cs, gr = random_sampling.sample_data(data=data, repeats=repeats, sample_size=sample_size)
                ax.plot(cs, gr, color='black', alpha=0.5, marker='.')
            title = random_sampling.get_plot_title(runs=runs, repeats=repeats, sample_size=sample_size)
            random_sampling.format_plot(fig, ax, f'{ws_name}, {title}', smoothing)
            plt.show()
        except Exception as e:
            print(f'Could not analyse {ws_name}:', e)
            plt.close()


def load_clean_data(ws, n: int) -> pd.DataFrame:
    """Returns a formatted pandas DataFrame, assuming data to be properly formatted"""
    return pd.DataFrame({
        'CS1': [float(c.value) for c in ws['A'][n:]],
        'GR2': [float(c.value) for c in ws['B'][n:]]
    })


if __name__ == '__main__':
    main(**SETTINGS)
