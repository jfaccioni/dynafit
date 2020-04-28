"""test_validator.py - unit tests for validator.py."""

import unittest
from random import randint
from string import ascii_uppercase
from unittest.mock import patch

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from src.exceptions import (BadCellStringError, DifferentSizeError, EmptyCellError, MismatchedColumnsError,
                            MismatchedRowsError, NoExcelFileError)
from src.validator import ExcelValidator

vec_isinstance = np.vectorize(isinstance)


class TestValidatorModule(unittest.TestCase):
    """Tests the validator.py module."""
    test_case_path = 'test/test_cases/excel_validator_test_case.xlsx'
    valid_cell_strings = ['A1', 'D78', 'C14', 'AB3', 'XYZ3']
    invalid_cell_strings = ['1', 'X', 'A1A', '4C', 'ABBA0303', '  E0  ', '', 'SOME OTHER THINGS', '?', ";;;;;;"]
    valid_cell_ranges = [('A1', 'A100'), ('B2', 'B3'), ('XX100', 'XX200')]
    invalid_cell_range_strings = [(a, b) for a, b in zip(invalid_cell_strings[::2], invalid_cell_strings[1::2])]
    invalid_cell_range_columns = [('A1', 'B100'), ('B2', 'C3'), ('XX100', 'XY200')]
    invalid_cell_range_rows = [('A100', 'A1'), ('B3', 'B2'), ('XX200', 'XX100')]

    def setUp(self) -> None:
        """Sets up each unit test by refreshing the ExcelValidator instance."""
        self.ev = ExcelValidator(workbook=openpyxl.Workbook(), sheetname='sheetname', cs_start_cell='A1',
                                 cs_end_cell='A10', gr_start_cell='B1', gr_end_cell='B10')

    def load_test_case(self, sheetname: str):
        """Loads the test case file as if it were passed in the ExcelValidator's __init__ method."""
        self.ev.wb = openpyxl.load_workbook(self.test_case_path)
        self.ev.sheetname = sheetname
        self.ev.ranges = {'CS': ['A1', ''], 'GR': ['B1', '']}

    def tearDown(self) -> None:
        """Deletes the ExcelValidator instance after a unit test ends."""
        del self.ev

    def test_none_value_as_workbook_raises_exception_when_no_data_was_loaded(self) -> None:
        with self.assertRaises(NoExcelFileError):
            ExcelValidator(workbook=None, sheetname='sheetname', cs_start_cell='A1', cs_end_cell='A10',
                           gr_start_cell='B1', gr_end_cell='B10')

    def test_ws_attribute_returns_excel_worksheet(self) -> None:
        self.load_test_case(sheetname='okay_spreadsheet')
        self.assertIsInstance(self.ev.ws, Worksheet)

    def test_validation_routine_passes_with_valid_cell_ranges(self) -> None:
        for start, end in self.valid_cell_ranges:
            with self.subTest(start=start, end=end):
                self.ev.ranges = {'CS': [start, end], 'GR': [start, end]}
                try:
                    self.ev.validation_routine()
                except (BadCellStringError, EmptyCellError, MismatchedColumnsError, MismatchedRowsError):
                    self.fail(f'Test failed: string {start}:{end} was intended to be a valid Excel cell.')

    def test_validation_routine_fails_with_invalid_cell_range_strings(self) -> None:
        for start, end in self.invalid_cell_range_strings:
            expected_exception = BadCellStringError if start else EmptyCellError
            with self.subTest(start=start, end=end), self.assertRaises(expected_exception):
                self.ev.ranges = {'CS': [start, end], 'GR': [start, end]}
                self.ev.validation_routine()

    def test_validation_routine_fails_with_invalid_cell_range_columns(self) -> None:
        for start, end in self.invalid_cell_range_columns:
            with self.subTest(start=start, end=end), self.assertRaises(MismatchedColumnsError):
                self.ev.ranges = {'CS': [start, end], 'GR': [start, end]}
                self.ev.validation_routine()

    def test_validation_routine_fails_with_invalid_cell_range_rows(self) -> None:
        for start, end in self.invalid_cell_range_rows:
            with self.subTest(start=start, end=end), self.assertRaises(MismatchedRowsError):
                self.ev.ranges = {'CS': [start, end], 'GR': [start, end]}
                self.ev.validation_routine()

    def test_convert_column_to_first_cell_adds_one_to_letter_columns(self) -> None:
        for letter in ascii_uppercase:
            with self.subTest(letter=letter):
                expected = letter + '1'
                actual = self.ev.convert_column_to_first_cell(cell_str=letter)
                self.assertEqual(expected, actual)

    def test_convert_column_to_first_cell_do_not_change_cell_name(self) -> None:
        for letter in ascii_uppercase:
            some_number = str(randint(1, 100))
            with self.subTest(letter=letter, number=some_number):
                expected = letter + some_number
                actual = self.ev.convert_column_to_first_cell(cell_str=letter + some_number)
                self.assertEqual(expected, actual)

    def test_convert_column_to_first_cell_edge_cases(self) -> None:
        # empty string remains the same
        expected = ''
        actual = self.ev.convert_column_to_first_cell(cell_str='')
        self.assertEqual(expected, actual)
        # Strings with 0 as the row number are dealt by downstream code, not here
        expected = 'A00000'
        actual = self.ev.convert_column_to_first_cell(cell_str='A00000')
        self.assertEqual(expected, actual)
        # Invalid strings are dealt by downstream code, not here
        expected = 'SOME REALLY BAD STRING1'
        actual = self.ev.convert_column_to_first_cell(cell_str='SOME REALLY BAD STRING')
        self.assertEqual(expected, actual)
        expected = 'ABBA1'
        actual = self.ev.convert_column_to_first_cell(cell_str='ABBA')
        self.assertEqual(expected, actual)

    def test_validate_cell_string_valid_strings_raise_no_error(self) -> None:
        for valid_cell_string in self.valid_cell_strings:
            with self.subTest(valid_cell_string=valid_cell_string):
                try:
                    self.ev.validate_cell_string(cell_str=valid_cell_string)
                except (BadCellStringError, EmptyCellError):
                    self.fail(f'Test failed: string {valid_cell_string} was intended to be a valid Excel cell.')

    def test_validate_cell_string_invalid_strings_raise_error(self) -> None:
        for invalid_cell_string in self.invalid_cell_strings:
            expected_exception = BadCellStringError if invalid_cell_string else EmptyCellError
            with self.subTest(invalid_cell_string=invalid_cell_string), self.assertRaises(expected_exception):
                self.ev.validate_cell_string(cell_str=invalid_cell_string)

    def test_is_valid_excel_valid_strings_return_true(self) -> None:
        for valid_cell_string in self.valid_cell_strings:
            with self.subTest(valid_cell_string=valid_cell_string):
                result = self.ev.is_valid_excel(cell_str=valid_cell_string)
                self.assertTrue(result)

    def test_is_valid_excel_invalid_strings_return_false(self) -> None:
        for invalid_cell_string in self.invalid_cell_strings:
            with self.subTest(invalid_cell_string=invalid_cell_string):
                result = self.ev.is_valid_excel(cell_str=invalid_cell_string)
                self.assertFalse(result)

    def test_validate_cell_range_valid_ranges_raise_no_errors(self) -> None:
        for start, end in self.valid_cell_ranges:
            with self.subTest(start=start, end=end):
                try:
                    self.ev.validate_cell_range(start=start, end=end)
                except (BadCellStringError, EmptyCellError, MismatchedRowsError, MismatchedColumnsError):
                    self.fail(f'Test failed: string {start}:{end} was intended to be a valid Excel range.')

    def test_validate_cell_range_invalid_range_strings_raise_string_error(self) -> None:
        for start, end in self.invalid_cell_range_strings:
            expected_exception = BadCellStringError if start else EmptyCellError
            with self.subTest(start=start, end=end), self.assertRaises(expected_exception):
                self.ev.validate_cell_range(start=start, end=end)

    def test_validate_cell_range_invalid_range_columns_raise_mismatched_columns_error(self) -> None:
        for start, end in self.invalid_cell_range_columns:
            with self.subTest(start=start, end=end), self.assertRaises(MismatchedColumnsError):
                self.ev.validate_cell_range(start=start, end=end)

    def test_validate_cell_range_invalid_range_columns_raise_mismatched_rows_error(self) -> None:
        for start, end in self.invalid_cell_range_rows:
            with self.subTest(start=start, end=end), self.assertRaises(MismatchedRowsError):
                self.ev.validate_cell_range(start=start, end=end)

    def test_validate_cell_ranges_share_same_column_returns_true_with_same_column_cell_strings(self) -> None:
        for start, end in self.valid_cell_ranges:
            with self.subTest(start=start, end=end):
                result = self.ev.validate_cell_ranges_share_same_column(start=start, end=end)
                self.assertTrue(result)
        for start, end in self.invalid_cell_range_rows:  # mismatched rows should cause no issues here
            with self.subTest(start=start, end=end):
                result = self.ev.validate_cell_ranges_share_same_column(start=start, end=end)
                self.assertTrue(result)

    def test_validate_cell_ranges_share_same_column_returns_false_with_different_column_cell_strings(self) -> None:
        for start, end in self.invalid_cell_range_columns:
            with self.subTest(start=start, end=end):
                result = self.ev.validate_cell_ranges_share_same_column(start=start, end=end)
                self.assertFalse(result)

    def test_validate_end_cell_comes_after_start_cell_returns_true_with_correct_rows_cell_strings(self) -> None:
        for start, end in self.valid_cell_ranges:
            with self.subTest(start=start, end=end):
                result = self.ev.validate_end_cell_comes_after_start_cell(start=start, end=end)
                self.assertTrue(result)
        for start, end in self.invalid_cell_range_columns:  # mismatched columns should cause no issues here
            with self.subTest(start=start, end=end):
                result = self.ev.validate_end_cell_comes_after_start_cell(start=start, end=end)
                self.assertTrue(result)

    def test_validate_end_cell_comes_after_start_cell_returns_false_with_incorrect_rows_cell_strings(self) -> None:
        for start, end in self.invalid_cell_range_rows:
            with self.subTest(start=start, end=end):
                result = self.ev.validate_end_cell_comes_after_start_cell(start=start, end=end)
                self.assertFalse(result)

    def test_get_data_works_with_good_data(self) -> None:
        self.load_test_case(sheetname='okay_spreadsheet')
        data = self.ev.get_data()
        self.assertIsInstance(data, pd.DataFrame)  # get_data returned a pandas DataFrame
        self.assertTrue(all(np.issubdtype(dtype, np.number) for dtype in data.dtypes))  # All columns are numeric
        self.assertEqual(len(data), 10)  # All 10 rows were parsed

    def test_get_data_ignores_missing_data(self) -> None:
        self.load_test_case(sheetname='blank_cells_spreadsheet')
        data = self.ev.get_data()
        self.assertIsInstance(data, pd.DataFrame)  # get_data returned a pandas DataFrame
        self.assertTrue(all(np.issubdtype(dtype, np.number) for dtype in data.dtypes))  # All columns are numeric
        self.assertEqual(len(data), 8)  # 8 rows were parsed, 2 rows with blank cells were ignored

    def test_get_data_ignores_non_numeric_data(self) -> None:
        self.load_test_case(sheetname='non_numeric_spreadsheet')
        data = self.ev.get_data()
        self.assertIsInstance(data, pd.DataFrame)  # get_data returned a pandas DataFrame
        self.assertTrue(all(np.issubdtype(dtype, np.number) for dtype in data.dtypes))  # All columns are numeric
        self.assertEqual(len(data), 8)  # 8 rows were parsed, 2 rows with non-numeric cells were ignored

    def test_get_data_raises_exception_different_sized_ranges(self) -> None:
        self.load_test_case(sheetname='okay_spreadsheet')
        self.ev.ranges['CS'][0] = 'A2'  # CS starts on A2, GR starts on B1, sizes will mismatch
        with self.assertRaises(DifferentSizeError):
            self.ev.get_data()

    def test_get_cell_values_extract_values_from_worksheet_column(self) -> None:
        self.load_test_case(sheetname='okay_spreadsheet')
        cell_range = 'A1:A11'
        cell_rows = self.ev.ws[cell_range]
        values = self.ev.get_cell_values(cell_rows)
        self.assertIsInstance(values, np.ndarray)  # get_cell_values returned a numpy array
        self.assertTrue(vec_isinstance(values, float).all())  # All values are numeric
        self.assertEqual(len(values), 11)  # All 11 rows were parsed, including header

    def test_get_cell_values_extract_values_coerces_headers_to_NaN(self) -> None:
        self.load_test_case(sheetname='okay_spreadsheet')
        cell_range = 'A1:A11'
        cell_rows = self.ev.ws[cell_range]
        values = self.ev.get_cell_values(cell_rows)
        self.assertTrue(np.isnan(values[0]))  # First row (header) was coerced to NaN
        self.assertTrue(~np.isnan(values[1:]).all())  # All other rows were not coerced to NaN

    def test_get_cell_values_extract_values_coerces_empty_cells_to_NaN(self) -> None:
        self.load_test_case(sheetname='blank_cells_spreadsheet')
        cell_range = 'A1:A11'
        cell_rows = self.ev.ws[cell_range]
        values = self.ev.get_cell_values(cell_rows)
        self.assertTrue(np.isnan(values[2]))  # Third row (empty cell) was coerced to NaN

    def test_get_cell_values_extract_values_coerces_non_numeric_to_NaN(self) -> None:
        self.load_test_case(sheetname='non_numeric_spreadsheet')
        cell_range = 'A1:A11'
        cell_rows = self.ev.ws[cell_range]
        values = self.ev.get_cell_values(cell_rows)
        self.assertTrue(np.isnan(values[2]))  # Third row (string 'hello') was coerced to NaN

    @patch('src.validator.ExcelValidator.ws')
    def test_get_end_cell_uses_max_row_attribute_of_ws(self, mock_ws) -> None:
        for max_row in range(20):
            with self.subTest(max_row=max_row):
                mock_ws.max_row = max_row
                start = 'A1'
                end = self.ev.get_end_cell(start=start)
                self.assertEqual(end, f'A{max_row}')

    def test_extract_letters_returns_all_letters_in_strings(self) -> None:
        test_case = 'A1'
        result = self.ev.extract_letters(test_case)
        self.assertEqual(result, 'A')
        test_case = 'I am a test string!'
        result = self.ev.extract_letters(test_case)
        self.assertEqual(result, 'Iamateststring')
        test_case = '10 > 100?'
        result = self.ev.extract_letters(test_case)
        self.assertEqual(result, '')

    def test_extract_digits_returns_all_digits_in_strings(self) -> None:
        test_case = 'A1'
        result = self.ev.extract_digits(test_case)
        self.assertEqual(result, '1')
        test_case = 'I am a test string!'
        result = self.ev.extract_digits(test_case)
        self.assertEqual(result, '')
        test_case = '10 > 100?'
        result = self.ev.extract_digits(test_case)
        self.assertEqual(result, '10100')


if __name__ == '__main__':
    unittest.main()
