"""test_interface.py - unit tests for interface.py."""

import unittest
from queue import Queue
from typing import List, Tuple
from unittest.mock import MagicMock, call, patch

import numpy as np
import openpyxl
import pandas as pd
from PySide2.QtCore import Qt, SIGNAL
from PySide2.QtTest import QTest
from PySide2.QtWidgets import QApplication, QFileDialog, QMessageBox, QTableWidgetItem
from matplotlib.pyplot import Axes

from src.core import dynafit
from src.exceptions import AbortedByUser
from src.interface import DynaFitGUI
from src.worker import Worker
from test.resources import SETTINGS_SCHEMA


class TestInterfaceModule(unittest.TestCase):
    """Tests the interface.py module."""
    test_case_path = 'test/test_cases/interface_test_case.xlsx'
    test_case_sheetnames = ['CS_GR', 'CS1_CS2']
    corrupted_test_case_path = 'test/test_cases/corrupted_test_case.xlsx'
    df_without_nans_before_filtering = pd.DataFrame({
            'col1': [0.1, 0.4, 0.7],
            'col2': [0.2, 0.5, 0.8],
        })
    df_without_nans_after_filtering = pd.DataFrame({
            'col1': ['0.1', '0.4', '0.7'],
            'col2': ['0.2', '0.5', '0.8'],
        })
    df_with_nans_before_filtering = pd.DataFrame({
            'col1': ['nan', 'some',    'text'],
            'col2': [0.2,   np.nan, 0.8],
        })
    df_with_nans_after_filtering = pd.DataFrame({
            'col1': ['', 'some',  'text'],
            'col2': ['0.2', '',  '0.8'],
        })

    @classmethod
    def setUpClass(cls) -> None:
        """Sets up the test suite by instantiating a QApplication."""
        cls.app = QApplication()

    def setUp(self) -> None:
        """Sets up each unit test by refreshing the DynaFitGUI instance."""
        self.ui = DynaFitGUI()

    def tearDown(self) -> None:
        """Deletes the DynaFitGUI instance after a unit test ends."""
        self.ui.destroy()

    def load_example_data(self) -> None:
        """Loads the test case file as if it were loaded through the DynaFitGUI."""
        self.ui.workbook = openpyxl.load_workbook(self.test_case_path)

    @property
    def mock_results(self) -> Tuple[MagicMock, MagicMock, MagicMock]:
        """Returns mocked results for the DynaFit Analysis"""
        params = MagicMock()
        plotter = MagicMock()
        plotter.plotting_methods = [plotter.plot_cvp_ax, plotter.plot_hypothesis_ax, plotter.plot_histogram_ax]
        df = MagicMock()
        return params, plotter, df

    @property
    def mock_dataframe(self) -> MagicMock:
        """Returns a mock dataframe for the DynaFit Analysis"""
        return self.mock_results[-1]

    @property
    def axes(self) -> List[Axes]:
        """Convenience property for accessing all Axes instances of the DynaFitGUI."""
        return [self.ui.cvp_ax, self.ui.hypothesis_ax, self.ui.histogram_ax]

    def set_results_table_names(self) -> None:
        """Adds some text to the results table so that the excel/csv buttons are able to read a placeholder text."""
        self.ui.results_table.setRowCount(2)
        self.ui.results_table.setItem(0, 1, QTableWidgetItem('filename'))  # noqa
        self.ui.results_table.setItem(1, 1, QTableWidgetItem('sheetname'))  # noqa

    def enable_export_buttons(self) -> None:
        """Enable buttons related to exporting the data. Simulates a finished DynaFit analysis."""
        self.ui.to_excel_button.setEnabled(True)
        self.ui.to_csv_button.setEnabled(True)

    def test_click_load_data_button_opens_load_data_dialog(self) -> None:
        QFileDialog.getOpenFileName = MagicMock(return_value=('', None))
        QTest.mouseClick(self.ui.load_data_button, Qt.LeftButton)  # noqa
        QFileDialog.getOpenFileName.assert_called()  # noqa

    @patch('src.interface.DynaFitGUI.load_data')
    def test_selecting_a_file_on_data_dialog_calls_load_data_method(self, mock_load_data) -> None:
        QFileDialog.getOpenFileName = MagicMock(return_value=('query', None))
        self.ui.load_data_dialog()
        QFileDialog.getOpenFileName.assert_called()  # noqa
        mock_load_data.assert_called_with(query='query')

    @patch('src.interface.DynaFitGUI.load_data')
    def test_not_selecting_a_file_on_data_dialog_do_not_call_load_data_method(self, mock_load_data) -> None:
        QFileDialog.getOpenFileName = MagicMock(return_value=('', None))
        self.ui.load_data_dialog()
        QFileDialog.getOpenFileName.assert_called()  # noqa
        mock_load_data.assert_not_called()

    @patch('src.interface.DynaFitGUI.load_data_success')
    @patch('src.interface.DynaFitGUI.raise_main_thread_error')
    def test_load_data_calls_load_data_success_when_file_is_not_corrupted(self, mock_raise_main_thread_error,
                                                                          mock_load_data_success) -> None:
        self.ui.load_data(query=self.test_case_path)
        mock_load_data_success.assert_called()
        mock_raise_main_thread_error.assert_not_called()

    @patch('src.interface.DynaFitGUI.load_data_success')
    @patch('src.interface.DynaFitGUI.raise_main_thread_error')
    def test_load_data_raises_exception_on_corrupted_excel_file(self, mock_raise_main_thread_error,
                                                                mock_load_data_success) -> None:
        self.ui.load_data(query=self.corrupted_test_case_path)
        mock_load_data_success.assert_not_called()
        mock_raise_main_thread_error.assert_called()

    def test_load_data_success_changes_ui_elements(self) -> None:
        self.load_example_data()
        # before load_data_success
        self.assertEqual(self.ui.input_filename_label.text(), '')
        self.assertFalse(self.ui.input_sheetname_combobox.isEnabled())
        self.assertEqual(self.ui.input_sheetname_combobox.count(), 0)
        self.ui.load_data_success(filename='filename')
        # after load_data_success
        self.assertEqual(self.ui.input_filename_label.text(), 'filename')
        self.assertTrue(self.ui.input_sheetname_combobox.isEnabled())
        self.assertEqual(self.ui.input_sheetname_combobox.count(), 2)
        for i, name in enumerate(self.test_case_sheetnames):
            self.ui.input_sheetname_combobox.setCurrentIndex(i)
            self.assertEqual(self.ui.input_sheetname_combobox.currentText(), name)

    def test_cs1_cs2_button_clicked_changes_ui_elements(self) -> None:
        QTest.mouseClick(self.ui.cs1_cs2_button, Qt.LeftButton)  # noqa
        self.assertTrue(self.ui.time_interval_label.isEnabled())
        self.assertTrue(self.ui.time_interval_spinbox.isEnabled())
        self.assertEqual(self.ui.CS_label.text(), 'Initial colony size column')
        self.assertEqual(self.ui.GR_label.text(), 'Final colony size column')

    def test_cs_cs_button_clicked_changes_ui_elements(self) -> None:
        QTest.mouseClick(self.ui.cs_gr_button, Qt.LeftButton)  # noqa
        self.assertFalse(self.ui.time_interval_label.isEnabled())
        self.assertFalse(self.ui.time_interval_spinbox.isEnabled())
        self.assertEqual(self.ui.CS_label.text(), 'Colony size column')
        self.assertEqual(self.ui.GR_label.text(), 'Growth rate column')

    def test_conf_int_checkbox_checked_changes_ui_elements(self) -> None:
        self.assertFalse(self.ui.conf_int_spinbox.isEnabled())
        QTest.mouseClick(self.ui.conf_int_checkbox, Qt.LeftButton)  # noqa
        self.assertTrue(self.ui.conf_int_spinbox.isEnabled())
        QTest.mouseClick(self.ui.conf_int_checkbox, Qt.LeftButton)  # noqa
        self.assertFalse(self.ui.conf_int_spinbox.isEnabled())

    @patch('src.interface.DynaFitGUI.dynafit_setup_before_running')
    @patch('src.interface.DynaFitGUI.get_dynafit_settings')
    @patch('src.interface.DynaFitGUI.raise_main_thread_error')
    @patch('src.interface.DynaFitGUI.dynafit_worker_run')
    def test_dynafit_run_success_calls_appropriate_downstream_methods(self, mock_dynafit_worker_run,
                                                                      mock_raise_main_thread_error,
                                                                      mock_get_dynafit_settings,
                                                                      mock_dynafit_setup_before_running) -> None:
        self.ui.dynafit_run()
        mock_dynafit_setup_before_running.assert_called()
        mock_get_dynafit_settings.assert_called()
        mock_raise_main_thread_error.assert_not_called()
        mock_dynafit_worker_run.assert_called_with(dynafit_settings=mock_get_dynafit_settings.return_value)

    @patch('src.interface.DynaFitGUI.dynafit_setup_before_running', side_effect=Exception)
    @patch('src.interface.DynaFitGUI.get_dynafit_settings')
    @patch('src.interface.DynaFitGUI.raise_main_thread_error')
    @patch('src.interface.DynaFitGUI.dynafit_worker_run')
    def test_dynafit_run_setup_failure_calls_main_thread_error(self, mock_dynafit_worker_run,
                                                               mock_raise_main_thread_error, mock_get_dynafit_settings,
                                                               mock_dynafit_setup_before_running) -> None:
        self.ui.dynafit_run()
        mock_dynafit_setup_before_running.assert_called()
        mock_raise_main_thread_error.assert_called()
        mock_get_dynafit_settings.assert_not_called()
        mock_dynafit_worker_run.assert_not_called()

    @patch('src.interface.DynaFitGUI.dynafit_setup_before_running')
    @patch('src.interface.DynaFitGUI.get_dynafit_settings', side_effect=Exception)
    @patch('src.interface.DynaFitGUI.raise_main_thread_error')
    @patch('src.interface.DynaFitGUI.dynafit_worker_run')
    def test_dynafit_run_get_settings_failure_calls_main_thread_error(self, mock_dynafit_worker_run,
                                                                      mock_raise_main_thread_error,
                                                                      mock_get_dynafit_settings,
                                                                      mock_dynafit_setup_before_running) -> None:
        self.ui.dynafit_run()
        mock_dynafit_setup_before_running.assert_called()
        mock_get_dynafit_settings.assert_called()
        mock_raise_main_thread_error.assert_called()
        mock_dynafit_worker_run.assert_not_called()

    def test_get_dynafit_settings_keys_are_dynafit_kwargs(self) -> None:
        settings = self.ui.get_dynafit_settings()
        for expected_argument_name in settings.keys():
            with self.subTest(expected_argument_name=expected_argument_name):
                self.assertIn(expected_argument_name, dynafit.__code__.co_varnames)

    def test_get_dynafit_settings_values_have_correct_types(self) -> None:
        settings = self.ui.get_dynafit_settings()
        del settings['workbook']  # do not test data (next tests do this)
        for expected_class, actual_value in zip(SETTINGS_SCHEMA.values(), settings.values()):
            with self.subTest(expected_class=expected_class, actual_value=actual_value):
                self.assertIsInstance(actual_value, expected_class)

    def test_get_dynafit_settings_no_data_loaded(self) -> None:
        settings = self.ui.get_dynafit_settings()
        self.assertIsNone(settings['workbook'])

    def test_get_dynafit_settings_data_loaded(self) -> None:
        self.load_example_data()
        settings = self.ui.get_dynafit_settings()
        self.assertIsInstance(settings['workbook'], openpyxl.Workbook)

    @patch('src.interface.dynafit')
    @patch('src.interface.QThreadPool.start')
    @patch('src.interface.Worker')
    def test_dynafit_worker_run_calls_worker_and_passes_it_to_threadpool(self, mock_worker, mock_threadpool_start,
                                                                         mock_dynafit) -> None:
        dynafit_settings = dict(a=1)
        self.ui.dynafit_worker_run(dynafit_settings=dynafit_settings)
        mock_worker.assert_called_with(func=mock_dynafit, a=1)
        mock_threadpool_start.assert_called_with(mock_worker.return_value)

    def test_dynafit_connect_worker_connects_worker_slots_to_interface_signals(self) -> None:
        signals = [SIGNAL('progress(int)'), SIGNAL('warning(PyObject)'), SIGNAL('finished()'),
                   SIGNAL('success(PyObject)'), SIGNAL('error(PyObject)')]
        w = Worker(func=print)
        for signal in signals:
            with self.subTest(signal=signal):
                self.assertEqual(w.signals.receivers(signal), 0)
        self.ui.connect_worker(worker=w)
        for signal in signals:
            with self.subTest(signal=signal):
                self.assertEqual(w.signals.receivers(signal), 1)

    def test_dynafit_setup_before_running_clears_axes(self) -> None:
        for ax in self.axes:
            ax.plot([1, 2], [3, 4])
            with self.subTest(ax=ax):
                self.assertTrue(ax.lines)
        self.ui.dynafit_setup_before_running()
        for ax in self.axes:
            with self.subTest(ax=ax):
                self.assertFalse(ax.lines)

    def test_dynafit_setup_before_running_turns_progress_bar_widgets_on(self) -> None:
        for widget in (self.ui.progress_bar, self.ui.progress_bar_label):
            with self.subTest(widget=widget):
                self.assertTrue(widget.isHidden())
        self.ui.dynafit_setup_before_running()
        for widget in (self.ui.progress_bar, self.ui.progress_bar_label):
            with self.subTest(widget=widget):
                self.assertFalse(widget.isHidden())

    def test_dynafit_setup_before_running_disables_plot_and_export_buttons(self) -> None:
        self.enable_export_buttons()
        for button in (self.ui.plot_button, self.ui.to_excel_button, self.ui.to_csv_button):
            with self.subTest(button=button):
                self.assertTrue(button.isEnabled())
        self.ui.dynafit_setup_before_running()
        for button in (self.ui.plot_button, self.ui.to_excel_button, self.ui.to_csv_button):
            with self.subTest(button=button):
                self.assertFalse(button.isEnabled())

    def test_dynafit_worker_progress_updated_sets_value_on_progress_bar(self) -> None:
        self.ui.dynafit_worker_progress_updated(10)
        self.assertEqual(self.ui.progress_bar.value(), 10)

    @patch('src.interface.QMessageBox')
    def test_dynafit_worker_small_sample_warning_displays_message_box(self, mock_message_box) -> None:
        warning_contents = (Queue(), {1: 2})
        self.ui.dynafit_worker_small_sample_size_warning(warning=warning_contents)
        mock_message_box.assert_called()
        mock_message_box.return_value.exec_.assert_called()

    def test_dynafit_worker_small_sample_warning_yes_button_puts_false_in_queue(self) -> None:
        q = Queue()
        warning_contents = (q, {1: 2})
        with patch('src.interface.QMessageBox.exec_', return_value=QMessageBox.Yes):
            self.ui.dynafit_worker_small_sample_size_warning(warning=warning_contents)
        self.assertFalse(q.get())

    def test_dynafit_worker_small_sample_warning_no_button_puts_true_in_queue(self) -> None:
        q = Queue()
        warning_contents = (q, {1: 2})
        with patch('src.interface.QMessageBox.exec_', return_value=QMessageBox.No):
            self.ui.dynafit_worker_small_sample_size_warning(warning=warning_contents)
        self.assertTrue(q.get())

    def test_dynafit_worker_raised_exception_removes_figure_title(self) -> None:
        self.ui.fig.suptitle('SOMETHING HERE')
        with patch('src.interface.DynaFitGUI.raise_worker_thread_error'):
            self.ui.dynafit_worker_raised_exception((Exception(), 'string'))
        self.assertEqual(self.ui.fig._suptitle.get_text(), '')  # call to private method

    def test_dynafit_worker_raised_exception_removes_figure_visibility(self) -> None:
        for ax in self.axes:
            ax.set_visible(True)
        with patch('src.interface.DynaFitGUI.raise_worker_thread_error'):
            self.ui.dynafit_worker_raised_exception((Exception(), 'string'))
        for ax in self.axes:
            with self.subTest(ax=ax):
                self.assertFalse(ax.get_visible())

    def test_dynafit_worker_raised_exception_clears_results_table(self) -> None:
        self.ui.results_table.clearContents = MagicMock()
        self.ui.results_table.clearContents.assert_not_called()
        with patch('src.interface.DynaFitGUI.raise_worker_thread_error'):
            self.ui.dynafit_worker_raised_exception((Exception(), 'string'))
        self.ui.results_table.clearContents.assert_called()

    def test_dynafit_worker_raised_exception_sets_result_dataframe_to_none(self) -> None:
        self.ui.results_dataframe = 'results'
        with patch('src.interface.DynaFitGUI.raise_worker_thread_error'):
            self.ui.dynafit_worker_raised_exception((Exception(), 'string'))
        self.assertIsNone(self.ui.results_dataframe)

    def test_dynafit_worker_raised_exception_passes_exception_to_handler_function(self) -> None:
        e = Exception()
        s = 'string'
        with patch('src.interface.DynaFitGUI.raise_worker_thread_error') as mock_worker_exception_handler:
            self.ui.dynafit_worker_raised_exception((e, s))
        mock_worker_exception_handler.assert_called_once_with((e, s))

    @patch('src.interface.DynaFitGUI.raise_worker_thread_error')
    def test_dynafit_worker_raised_exception_no_exception_shown_when_aborted_by_user(self, mock_thread_error) -> None:
        mock_thread_error.assert_not_called()
        self.ui.dynafit_worker_raised_exception((AbortedByUser(), 'string'))
        mock_thread_error.assert_not_called()

    def test_dynafit_worker_raised_no_exception_reenables_gui_buttons(self) -> None:
        self.assertFalse(self.ui.to_excel_button.isEnabled())
        self.assertFalse(self.ui.to_csv_button.isEnabled())
        self.ui.dynafit_worker_raised_no_exceptions(results=self.mock_results)
        self.assertTrue(self.ui.to_excel_button.isEnabled())
        self.assertTrue(self.ui.to_csv_button.isEnabled())

    def test_dynafit_worker_raised_no_exception_sets_all_axes_visible(self) -> None:
        for ax in self.axes:
            with self.subTest(ax=ax):
                self.assertFalse(ax.get_visible())
        self.ui.dynafit_worker_raised_no_exceptions(results=self.mock_results)
        for ax in self.axes:
            with self.subTest(ax=ax):
                self.assertTrue(ax.get_visible())

    def test_dynafit_worker_raised_no_exception_sets_dataframe_results_attribute(self) -> None:
        results = self.mock_results
        *_, df = results
        self.assertIsNone(self.ui.results_dataframe)
        self.ui.dynafit_worker_raised_no_exceptions(results=results)
        self.assertIsNotNone(self.ui.results_dataframe)

    @patch('src.interface.DynaFitGUI.remove_nan_strings')
    def test_dynafit_worker_raised_no_exception_calls_remove_nan_strings(self, mock_remove_nan_strings) -> None:
        results = self.mock_results
        *_, df = results
        self.ui.dynafit_worker_raised_no_exceptions(results=results)
        mock_remove_nan_strings.assert_called_with(df=df)

    def test_dynafit_worker_raised_no_exception_calls_plotter_methods_with_correct_axes(self) -> None:
        results = self.mock_results
        _, plotter, _ = results
        for method in plotter.plotting_methods:
            method.assert_not_called()
        self.ui.dynafit_worker_raised_no_exceptions(results=results)
        for method, ax in zip(plotter.plotting_methods, self.axes):
            if ax == self.ui.hypothesis_ax:
                method.assert_called_once_with(ax=ax, xlims=self.ui.cvp_ax.get_xlim())
            else:
                method.assert_called_once_with(ax=ax)

    def test_dynafit_worker_raised_no_exception_calls_sets_figure_title_with_correct_parameters(self) -> None:
        results = self.mock_results
        params, *_ = results
        params.__getitem__.assert_not_called()
        self.ui.dynafit_worker_raised_no_exceptions(results=results)
        params.__getitem__.assert_has_calls([call('filename'), call('sheetname')], any_order=True)

    @patch('src.interface.DynaFitGUI.set_results_table')
    def test_dynafit_worker_raised_no_exception_calls_sets_results_table(self, mock_set_results_table) -> None:
        mock_set_results_table.assert_not_called()
        self.ui.dynafit_worker_raised_no_exceptions(results=self.mock_results)
        mock_set_results_table.assert_called_once()

    def test_set_figure_title_sets_the_figure_title(self) -> None:
        self.assertIsNone(self.ui.fig._suptitle)
        self.ui.set_figure_title(filename='filename', sheetname='sheetname')
        for param in ['filename', 'sheetname']:
            self.assertIn(param, self.ui.fig._suptitle.get_text())  # call to private method

    def test_set_results_table_clears_table_contents(self) -> None:
        self.ui.results_table.clearContents = MagicMock()
        self.ui.results_table.clearContents.assert_not_called()
        self.ui.set_results_table(self.mock_dataframe)
        self.ui.results_table.clearContents.assert_called_once()

    def test_set_results_table_sets_table_number_of_rows(self) -> None:
        self.ui.results_table.setRowCount = MagicMock()
        self.ui.results_table.setRowCount.assert_not_called()
        self.ui.set_results_table(self.mock_dataframe)
        self.ui.results_table.setRowCount.assert_called_once_with(len(self.mock_dataframe))

    def test_set_results_table_adds_rows_to_table(self) -> None:
        self.assertEqual(self.ui.results_table.rowCount(), 0)
        self.ui.set_results_table(self.df_without_nans_after_filtering)
        self.assertEqual(self.ui.results_table.rowCount(), 3)

    def test_set_results_table_adds_items_to_table_and_replaces_nan_with_empty_strings(self) -> None:
        for df_before, df_after in zip([self.df_without_nans_before_filtering, self.df_with_nans_before_filtering],
                                       [self.df_without_nans_after_filtering, self.df_with_nans_after_filtering]):
            self.ui.results_table.clearContents()
            self.ui.set_results_table(df_before)
            for row in range(self.ui.results_table.rowCount()):
                for column in range(2):
                    expected_value = df_after.iloc[row, column]
                    actual_value = self.ui.results_table.item(row, column).text()
                    with self.subTest(row=row, column=column, expected_value=expected_value, actual_value=actual_value,
                                      df_before=df_before, df_after=df_after):
                        self.assertEqual(expected_value, actual_value)

    def test_remove_nan_strings_returns_dataframe_with_no_nan_strings(self) -> None:
        df = pd.DataFrame({
            'col1': ['nan', 'something', 'else'],
            'col2': [0.2, np.nan, 0.8],
        })
        expected_df = pd.DataFrame({
            'col1': ['', 'something', 'else'],
            'col2': [0.2, np.nan, 0.8],
        })
        actual_df = self.ui.remove_nan_strings(df=df)
        pd.testing.assert_frame_equal(expected_df, actual_df)

    def test_dynafit_worker_has_finished_enables_plot_button(self) -> None:
        self.ui.plot_button.setText('Something else')
        self.ui.plot_button.setEnabled(False)
        self.ui.dynafit_worker_has_finished()
        self.assertEqual(self.ui.plot_button.text(), 'Plot CVP')
        self.assertTrue(self.ui.plot_button.isEnabled())

    def test_dynafit_worker_has_finished_hides_progress_bar_and_label(self) -> None:
        self.ui.progress_bar.setHidden(True)
        self.ui.progress_bar_label.setHidden(True)
        self.ui.dynafit_worker_has_finished()
        self.assertTrue(self.ui.progress_bar.isHidden())
        self.assertTrue(self.ui.progress_bar_label.isHidden())

    def test_dynafit_worker_has_finished_sets_progress_bar_to_zero(self) -> None:
        self.ui.progress_bar.setValue(10)
        self.ui.dynafit_worker_has_finished()
        self.assertEqual(self.ui.progress_bar.value(), 0)

    def test_dynafit_worker_has_finished_calls_canvas_draw_method(self) -> None:
        self.ui.canvas.draw = MagicMock()
        self.ui.canvas.draw.assert_not_called()
        self.ui.dynafit_worker_has_finished()
        self.ui.canvas.draw.assert_called()

    @patch('src.interface.DynaFitGUI.save_excel_dialog')
    def test_click_to_excel_button_calls_save_excel_dialog(self, mock_save_excel_dialog) -> None:
        self.enable_export_buttons()
        mock_save_excel_dialog.assert_not_called()
        QTest.mouseClick(self.ui.to_excel_button, Qt.LeftButton)  # noqa
        mock_save_excel_dialog.assert_called()

    @patch('src.interface.DynaFitGUI.save_excel')
    def test_save_excel_dialog_calls_save_excel_if_user_chooses_a_file(self, mock_save_excel) -> None:
        self.set_results_table_names()
        mock_save_excel.assert_not_called()
        with patch('PySide2.QtWidgets.QFileDialog.getSaveFileName', return_value=('filename', None)):
            self.ui.save_excel_dialog()
        mock_save_excel.assert_called_with(path='filename', sheet_name='filename_sheetname')

    @patch('src.interface.DynaFitGUI.save_excel')
    def test_save_excel_dialog_does_not_call_save_excel_if_user_does_not_choose_a_file(self, mock_save_excel) -> None:
        self.set_results_table_names()
        mock_save_excel.assert_not_called()
        with patch('PySide2.QtWidgets.QFileDialog.getSaveFileName', return_value=('', None)):
            self.ui.save_excel_dialog()
        mock_save_excel.assert_not_called()

    @patch('src.interface.pd.DataFrame.to_excel')
    def test_save_excel_adds_xlsx_extension_to_user_selected_file_name(self, mock_to_excel) -> None:
        self.ui.results_dataframe = pd.DataFrame()
        self.ui.save_excel(path='path', sheet_name='sheet_name')
        mock_to_excel.assert_called_with('path.xlsx', index=False, sheet_name='sheet_name')

    @patch('src.interface.DynaFitGUI.save_csv_dialog')
    def test_click_to_csv_button_calls_save_csv_dialog(self, mock_save_csv_dialog) -> None:
        self.enable_export_buttons()
        mock_save_csv_dialog.assert_not_called()
        QTest.mouseClick(self.ui.to_csv_button, Qt.LeftButton)  # noqa
        mock_save_csv_dialog.assert_called()

    @patch('src.interface.DynaFitGUI.save_csv')
    def test_save_csv_dialog_calls_save_csv_if_user_chooses_a_file(self, mock_save_csv) -> None:
        self.set_results_table_names()
        mock_save_csv.assert_not_called()
        with patch('PySide2.QtWidgets.QFileDialog.getSaveFileName', return_value=('filename', None)):
            self.ui.save_csv_dialog()
        mock_save_csv.assert_called_with(path='filename')

    @patch('src.interface.DynaFitGUI.save_csv')
    def test_save_csv_dialog_does_not_call_save_csv_if_user_does_not_choose_a_file(self, mock_save_csv) -> None:
        self.ui.results_dataframe = 'results'
        self.set_results_table_names()
        mock_save_csv.assert_not_called()
        with patch('PySide2.QtWidgets.QFileDialog.getSaveFileName', return_value=('', None)):
            self.ui.save_csv_dialog()
        mock_save_csv.assert_not_called()

    @patch('src.interface.pd.DataFrame.to_csv')
    def test_save_csv_adds_csv_extension_to_user_selected_file_name(self, mock_to_csv) -> None:
        self.ui.results_dataframe = pd.DataFrame()
        self.ui.save_csv(path='path')
        mock_to_csv.assert_called_with('path.csv', index=False)

    @patch('src.interface.DynaFitGUI.show_error_message')
    def test_main_thread_error_calls_show_error_message_with_traceback(self, mock_show_error_message) -> None:
        mock_show_error_message.assert_not_called()
        message = 'values not ok'
        error = ValueError(message)
        name = 'ValueError:\n' + message
        trace = 'NoneType: None\n'  # default trace when no Exceptions have been raised
        self.ui.raise_main_thread_error(error=error)
        mock_show_error_message.assert_called_with(name=name, trace=trace)

    @patch('src.interface.DynaFitGUI.show_error_message')
    def test_worker_thread_error_calls_show_error_message_with_string(self, mock_show_error_message) -> None:
        mock_show_error_message.assert_not_called()
        message = 'values not ok'
        error = ValueError(message)
        name = 'ValueError:\n' + message
        traceback_string = 'Some traceback message here'
        self.ui.raise_worker_thread_error(exception_tuple=(error, traceback_string))
        mock_show_error_message.assert_called_with(name=name, trace=traceback_string)

    @patch('src.interface.QMessageBox')
    def test_show_error_message_shows_message_box_with_error_contents(self, mock_message_box) -> None:
        self.ui.show_error_message(name='name', trace='trace')
        mock_message_box.assert_called_with(self.ui, windowTitle='An error occurred!', text='name',
                                            detailedText='trace')
        mock_message_box.return_value.exec_.assert_called()

    def test_debug(self) -> None:
        """No need to test debug method."""

    def test_resizeEvent(self) -> None:
        """No need to test resizeEvent method."""

    def test_eventFilter(self) -> None:
        """eventFilter method from SO (see source code for link)."""

    def test_copy_selection(self) -> None:
        """copy_selection method from SO (see source code for link)."""

    def test_main(self) -> None:
        """No need to test interface.main."""

    def test_main_debug(self) -> None:
        """No need to test interface.main_debug."""


if __name__ == '__main__':
    unittest.main()
