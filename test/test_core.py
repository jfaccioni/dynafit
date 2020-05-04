"""test_core.py - unit tests for core.py."""

import unittest
from unittest.mock import MagicMock, patch

import openpyxl
from PySide2.QtCore import SIGNAL

from src.core import *
from src.exceptions import TooManyGroupsError
from src.validator import ExcelValidator


class TestCoreModule(unittest.TestCase):
    """Tests the core.py module."""
    test_case_path = 'test/test_cases/core_test_case.xlsx'

    @classmethod
    def setUpClass(cls) -> None:
        """Sets up the the whole test suite by loading the test case Excel spreadsheet into memory."""
        cls.workbook = openpyxl.load_workbook(cls.test_case_path, data_only=True)
        cls.cs_gr_df = ExcelValidator(workbook=cls.workbook, sheetname='CS_GR', cs_start_cell='A1', cs_end_cell='',
                                      gr_start_cell='B1', gr_end_cell='').get_data()
        cls.cs1_cs2_df = ExcelValidator(workbook=cls.workbook, sheetname='CS1_CS2', cs_start_cell='A1', cs_end_cell='',
                                        gr_start_cell='B1', gr_end_cell='').get_data()

    @property
    def df(self) -> pd.DataFrame:
        """Alias for self.cs_gr_df, used in almost all tests involving the input data."""
        return self.cs_gr_df

    def get_test_case_calculated_gr(self) -> pd.Series:
        """Returns the GR calculated in the test case Excel."""
        ws = openpyxl.load_workbook(self.test_case_path, data_only=True)['GR_series']
        values = [c[0].value for c in ws[f'A2:A{ws.max_row}']]
        return pd.Series(values)

    @staticmethod
    def get_values_for_results_dataframe() -> Dict:
        """Convenience function to return a collection of elements used to mock a call to get_results_dataframe."""
        return {
            'original_parameters': {'parameter': 'value'},
            'xs': np.array([0]),
            'ys': np.array([0]),
            'cumulative_ys': np.array([0]),
            'endpoint_ys': np.array([0]),
            'show_ci': True,
            'upper_ys': np.array([0]),
            'lower_ys': np.array([0]),
            'cumulative_upper_ys': np.array([0]),
            'cumulative_lower_ys': np.array([0]),
            'endpoint_upper_ys': np.array([0]),
            'endpoint_lower_ys': np.array([0]),
        }

    def test_dynafit(self) -> None:
        pass  # TODO: missing test

    @patch('src.core.filter_colony_sizes_less_than_one')
    @patch('src.core.add_growth_rate_column')
    @patch('src.core.filter_outliers')
    def test_preprocess_data_calls_filter_colony_size_less_than_one(self, mock_filter_outliers,
                                                                    mock_add_growth_rate_column,
                                                                    mock_filter_colony_sizes_less_than_one) -> None:
        df = self.df.copy()
        preprocess_data(df=df, calculate_growth_rate=False, time_delta=24, remove_outliers=False)
        mock_filter_outliers.assert_not_called()
        mock_add_growth_rate_column.assert_not_called()
        mock_filter_colony_sizes_less_than_one.assert_called_with(df=df)

    @patch('src.core.filter_colony_sizes_less_than_one')
    @patch('src.core.add_growth_rate_column')
    @patch('src.core.filter_outliers')
    def test_preprocess_data_calls_calculate_growth_rate_conditionally(self, mock_filter_outliers,
                                                                       mock_add_growth_rate_column,
                                                                       mock_filter_colony_sizes_less_than_one) -> None:
        df = self.df.copy()
        preprocess_data(df=df, calculate_growth_rate=True, time_delta=24, remove_outliers=False)
        mock_filter_outliers.assert_not_called()
        mock_add_growth_rate_column.assert_called_with(df=mock_filter_colony_sizes_less_than_one.return_value,
                                                       time_delta=24)
        mock_filter_colony_sizes_less_than_one.assert_called_with(df=df)

    @patch('src.core.filter_colony_sizes_less_than_one')
    @patch('src.core.add_growth_rate_column')
    @patch('src.core.filter_outliers')
    def test_preprocess_data_calls_filter_outliers_conditionally(self, mock_filter_outliers,
                                                                 mock_add_growth_rate_column,
                                                                 mock_filter_colony_sizes_less_than_one) -> None:
        df = self.df.copy()
        preprocess_data(df=df, calculate_growth_rate=False, time_delta=24, remove_outliers=True)
        mock_filter_outliers.assert_called_with(df=mock_filter_colony_sizes_less_than_one.return_value)
        mock_add_growth_rate_column.assert_not_called()
        mock_filter_colony_sizes_less_than_one.assert_called_with(df=df)

    @patch('src.core.filter_colony_sizes_less_than_one')
    @patch('src.core.add_growth_rate_column')
    @patch('src.core.filter_outliers')
    def test_preprocess_data_calls_full_call_chain(self, mock_filter_outliers, mock_add_growth_rate_column,
                                                   mock_filter_colony_sizes_less_than_one) -> None:
        df = self.df.copy()
        preprocess_data(df=df, calculate_growth_rate=True, time_delta=24, remove_outliers=True)
        mock_filter_outliers.assert_called_with(df=mock_add_growth_rate_column.return_value)
        mock_add_growth_rate_column.assert_called_with(df=mock_filter_colony_sizes_less_than_one.return_value,
                                                       time_delta=24)
        mock_filter_colony_sizes_less_than_one.assert_called_with(df=df)

    def test_filter_colony_sizes_less_than_one_removes_colonies_with_less_than_one_cell(self) -> None:
        df = self.df.copy()
        df.CS -= 1
        length_before = len(df)
        self.assertTrue(any(cs < 1 for cs in df.CS))
        df = filter_colony_sizes_less_than_one(df=df)
        length_after = len(df)
        self.assertFalse(any(cs < 1 for cs in df.CS))
        self.assertGreater(length_before, length_after)

    def test_filter_colony_sizes_less_than_one_preserves_dataframe_with_colonies_with_at_least_one_cell(self) -> None:
        df = self.df.copy()
        length_before = len(df)
        self.assertFalse(any(cs < 1 for cs in df.CS))
        df = filter_colony_sizes_less_than_one(df=df)
        length_after = len(df)
        self.assertFalse(any(cs < 1 for cs in df.CS))
        self.assertEqual(length_before, length_after)

    def test_calculate_growth_rate_calculates_growth_rate_properly(self) -> None:
        expected_gr = self.get_test_case_calculated_gr()
        actual_gr = add_growth_rate_column(df=self.cs1_cs2_df, time_delta=72.0).GR
        np.testing.assert_allclose(expected_gr.values, actual_gr.values)

    def test_calculate_growth_rate_raises_value_error_on_negative_log(self) -> None:
        with self.assertRaises(ValueError):
            add_growth_rate_column(df=self.cs_gr_df, time_delta=72.0)

    def test_filter_outliers_remove_growth_rate_outliers_from_dataframe(self) -> None:
        data = pd.DataFrame({
            'GR': [1, 2, 3, 4, 50]  # 50 is the outlier
        })
        filtered_data = filter_outliers(data)
        self.assertEqual(len(data), len(filtered_data) + 1)

    def test_filter_outliers_does_not_remove_non_growth_rate_outliers_from_dataframe(self) -> None:
        data = pd.DataFrame({
            'GR': [1, 2, 3, 4, 5]  # no outliers here
        })
        filtered_data = filter_outliers(data)
        self.assertEqual(len(data), len(filtered_data))

    def test_divide_sample_into_groups_puts_small_colony_size_population_into_individual_groups(self) -> None:
        data = pd.DataFrame({
            'CS': [1, 1, 2, 2, 3, 3]
        })
        grouped_data = divide_sample_into_groups(data, individual_colonies=3, large_groups=0)
        expected_groups = np.array([1, 1, 2, 2, 3, 3])
        actual_groups = grouped_data['groups'].values
        np.testing.assert_allclose(expected_groups, actual_groups)

    def test_divide_sample_into_groups_puts_large_colony_size_population_into_collective_groups(self) -> None:
        data = pd.DataFrame({
            'CS': [10, 12, 24, 25, 31, 32]
        })
        grouped_data = divide_sample_into_groups(data, individual_colonies=0, large_groups=3)
        expected_groups = np.array([1, 1, 2, 2, 3, 3])
        actual_groups = grouped_data['groups'].values
        np.testing.assert_allclose(expected_groups, actual_groups)

    def test_divide_sample_into_groups_both_small_and_large_colonies_present(self) -> None:
        data = pd.DataFrame({
            'CS': [1, 1, 2, 2, 3, 3, 10, 12, 24, 25, 31, 32]
        })
        grouped_data = divide_sample_into_groups(data, individual_colonies=3, large_groups=3)
        expected_groups = np.array([1, 1, 2, 2, 3, 3, 4, 4, 5, 5, 6, 6])
        actual_groups = grouped_data['groups'].values
        np.testing.assert_allclose(expected_groups, actual_groups)

    def test_divide_sample_into_groups_raises_error_when_too_many_groups_are_present_for_large_colonies(self) -> None:
        data = pd.DataFrame({
            'CS': [10, 11, 10, 12, 15, 15]
        })
        with self.assertRaises(TooManyGroupsError):
            divide_sample_into_groups(data, individual_colonies=0, large_groups=6)

    def test_sample_size_warning_info_returns_dictionary_of_warnings(self) -> None:
        data = pd.DataFrame({
            'CS': [1, 1, 1, 2, 2, 2, 3],
            'groups': [1, 1, 1, 2, 2, 2, 3]
        })
        expected_warning_dictionary = {3: (1, 3.0)}  # group number: (number of instances, CS mean)
        actual_warning_dictionary = sample_size_warning_info(df=data, warning_level=2)
        self.assertEqual(expected_warning_dictionary, actual_warning_dictionary)

    def test_sample_size_warning_info_returns_none_when_dictionary_is_empty(self) -> None:
        data = pd.DataFrame({
            'CS': [1, 1, 1, 2, 2, 2, 3, 3, 3],
            'groups': [1, 1, 1, 2, 2, 2, 3, 3, 3]
        })
        return_value = sample_size_warning_info(df=data, warning_level=2)
        self.assertIsNone(return_value)

    def test_sample_size_warning_answer_returns_false_when_no_warning_dict_was_created(self) -> None:
        return_value = sample_size_warning_answer(warning_info=None, callback=SIGNAL('warning(PyObject)'))
        self.assertFalse(return_value)

    def test_sample_size_warning_answer_returns_false_when_queue_receives_false_value(self) -> None:
        with patch('src.core.Queue.get', return_value=False):
            return_value = sample_size_warning_answer(warning_info={3: 1}, callback=MagicMock())
        self.assertFalse(return_value)

    def test_sample_size_warning_answer_returns_true_when_queue_receives_true_value(self) -> None:
        with patch('src.core.Queue.get', return_value=True):
            return_value = sample_size_warning_answer(warning_info={3: 1}, callback=MagicMock())
        self.assertTrue(return_value)

    def test_get_original_sample_parameters_returns_mean_values_for_each_group(self) -> None:
        test_case_df = pd.DataFrame({
            'CS': [1, 2, 3, 4, 5, 6],
            'GR': [10, 20, 30, 40, 50, 60],
            'groups': [1, 1, 2, 2, 3, 3]
        })
        expected_xs = np.log2([1.5, 3.5, 5.5])  # CS mean for each group
        expected_ys = np.log2([50, 50, 50])  # GR var for each group
        actual_xs, actual_ys = get_original_sample_parameters(df=test_case_df)
        for expected_array, actual_array in zip([expected_xs, expected_ys], [actual_xs, actual_ys]):
            with self.subTest(expected_array=expected_array, actual_array=actual_array):
                np.testing.assert_allclose(expected_array, actual_array)

    def test_get_original_sample_statistics_returns_variance_and_its_se_for_each_group(self) -> None:
        pass  # TODO: missing test

    def test_get_histogram_values_gets_expected_values_properly(self) -> None:
        test_case_df = pd.DataFrame({
            'CS': [1, 2, 3, 4, 5, 6],
            'groups': [1, 1, 2, 2, 3, 3]
        })
        expected_xs = np.array([1, 2, 3, 4, 5, 6])  # same values as test_case_df['CS']
        expected_intervals = np.array([2, 4, 6])  # max value inside each group
        actual_xs, actual_intervals = get_histogram_values(df=test_case_df)
        for expected, actual in zip([expected_xs, expected_intervals], [actual_xs, actual_intervals]):
            with self.subTest(expected=expected, actual=actual):
                np.testing.assert_allclose(expected, actual)

    def test_bootstrap_data(self) -> None:
        pass  # TODO: missing test

    def test_get_bootstrap_params(self) -> None:
        pass  # TODO: missing test

    def test_get_t_stat(self) -> None:
        pass  # TODO: missing test

    def test_emit_bootstrap_progress_emits_progress_as_percentage(self) -> None:
        callback = MagicMock()
        emit_bootstrap_progress(current=1, total=100, callback=callback)
        percentage = 1
        callback.emit.assert_called_with(percentage)

    def test_postprocess_data(self) -> None:
        pass  # TODO: missing test

    def test_drop_inf_nan_values(self) -> None:
        pass  # TODO: missing test

    def test_add_log_columns_adds_log_columns_names_to_dataframe(self) -> None:
        test_case_df = pd.DataFrame({
            'CS_mean': [1, 2, 4, 8, 16],
            'GR_var': [32, 64, 128, 256, 512]
        })
        df_with_log_columns = add_log2_columns(df=test_case_df, column_names=test_case_df.columns)
        for column_name in test_case_df.columns:
            new_column_name = 'log2_' + column_name
            with self.subTest(new_column_name=new_column_name):
                self.assertIn(new_column_name, df_with_log_columns)

    def test_add_log_columns_properly_calculates_log_columns(self) -> None:
        expected_df = pd.DataFrame({
            'log2_CS_mean': [0, 1, 2, 3, 4],
            'log2_GR_var': [5, 6, 7, 8, 9]
        })
        actual_df = add_log2_columns(df=pd.DataFrame({
            'CS_mean': [1, 2, 4, 8, 16],
            'GR_var': [32, 64, 128, 256, 512]
        }), column_names=['CS_mean', 'GR_var'])
        for column_name in expected_df.columns:
            expected_values = expected_df[column_name].values
            actual_values = actual_df[column_name].values
            np.testing.assert_allclose(expected_values, actual_values)

    def test_get_get_scatter_values_returns_appropriate_scatter_values(self) -> None:
        test_case_df = pd.DataFrame({
            'log2_bootstrap_CS_mean': [1, 2, 3, 4, 5, 6],
            'log2_bootstrap_GR_var': [10, 20, 30, 40, 50, 60],
            'groups': [1, 1, 2, 2, 3, 3]
        })
        expected_xs = test_case_df['log2_bootstrap_CS_mean'].values
        expected_ys = test_case_df['log2_bootstrap_GR_var'].values
        expected_colors = np.array((['red'] * 4) + (['gray'] * 2))
        actual_values = get_bootstrap_scatter_values(df=test_case_df)
        for expected, actual in zip([expected_xs, expected_ys, expected_colors], actual_values):
            with self.subTest(expected=expected, actual=actual):
                np.testing.assert_array_equal(expected, actual)

    def test_get_get_violin_values_returns_appropriate_violin_values(self) -> None:
        test_case_df = pd.DataFrame({
            'log2_bootstrap_CS_mean': [10, 20, 30, 40, 50, 60],
            'log2_bootstrap_GR_var': [10, 20, 30, 40, 50, 60],
            'groups': [1, 1, 2, 2, 3, 3]
        })
        expected_xs = np.array([15, 35, 55])
        expected_ys = [np.array([10, 20]), np.array([30, 40]), np.array([50, 60])]
        actual_values = get_bootstrap_violin_values(df=test_case_df)
        for expected, actual in zip([expected_xs, expected_ys], actual_values):
            with self.subTest(expected=expected, actual=actual):
                np.testing.assert_array_equal(expected, actual)

    def test_get_bootstrap_violin_statistics(self) -> None:
        pass  # TODO: missing test

    def test_get_cumulative_hypothesis_values_calculates_areas_properly(self) -> None:
        # test case is a trapezium (area = 2) above the X axis, and a triangle below the X axis (area = -0.5)
        test_case_xs = np.array([0, 1, 2, 3, 4])
        test_case_ys = np.array([0, 1, 1, 0, -1])
        # partials for integrated area of trapezium above
        expected_area_array = np.array([0, 0.5, 1.5, 2, 1.5])
        # expected triangular area for line with angular coefficient = -1
        expected_triangle_array = np.array([0, -0.5, -2, -4.5, -8])
        expected_result_array = np.nan_to_num(expected_area_array / expected_triangle_array, posinf=0.0, neginf=0.0)
        actual_result_array = get_cumulative_hypothesis_values(xs=test_case_xs, ys=test_case_ys)
        np.testing.assert_allclose(expected_result_array, actual_result_array)

    def test_get_endpoint_hypothesis_values_calculates_distances_properly(self) -> None:
        # test case is a trapezium (area = 2) above the X axis, and a triangle below the X axis (area = -0.5)
        test_case_xs = np.array([0, 1, 2, 3, 4])
        test_case_ys = np.array([0, 1, 1, 0, -1])
        # expected result is divided by (-test_case_xs), since test_case_xs is a line with angular coefficient = 1
        expected_result_array = np.nan_to_num(test_case_ys / (-test_case_xs), posinf=0.0, neginf=0.0)
        actual_result_array = get_endpoint_hypothesis_values(xs=test_case_xs, ys=test_case_ys)
        np.testing.assert_allclose(expected_result_array, actual_result_array)

    def test_get_mean_line_ci_returns_two_numpy_arrays_with_one_element_for_each_group(self) -> None:
        test_case_df = pd.DataFrame({
            'log2_GR_var': [10, 20, 30, 40, 50, 60],
            'groups': [1, 1, 2, 2, 3, 3],
            't_stat': [1, 2, 3, 4, 5, 6]
        })
        upp, low = get_bootstrap_ci(df=test_case_df, confidence_value=0.95, original_var=np.array([1, 2, 3]),
                                    original_var_se=np.array([1, 2, 3]))
        number_of_groups = len(test_case_df['groups'].unique())
        for array in (upp, low):
            with self.subTest(array=array):
                self.assertIsInstance(array, np.ndarray)
                self.assertEqual(len(array), number_of_groups)

    def test_calculate_bootstrap_ci_from_t_distribution_calculates_ci_properly(self) -> None:
        t = pd.Series([0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100])  # q1 = 25.0, q3 = 75.0
        alpha = 0.5  # Gets q1 and q3 of t distribution
        sample_stat = 100
        sample_stat_se = 1
        expected_ci = (np.log2(100 - 25), np.log2(100 - 75))
        actual_ci = calculate_bootstrap_ci_from_t_distribution(t_distribution=t, alpha=alpha, sample_stat=sample_stat,
                                                               sample_stat_se=sample_stat_se)
        for expected_value, actual_value in zip(expected_ci, actual_ci):
            with self.subTest(expected_value=expected_value, actual_value=actual_value):
                self.assertEqual(expected_value, actual_value)

    def test_results_to_dataframe_returns_pandas_dataframe_of_data_passed_in(self) -> None:
        kwargs = self.get_values_for_results_dataframe()
        result = results_to_dataframe(**kwargs)
        self.assertIsInstance(result, pd.DataFrame)

    def test_results_to_dataframe_pads_dataframe_with_nans_when_data_has_different_sizes(self) -> None:
        kwargs = self.get_values_for_results_dataframe()
        initial_result = results_to_dataframe(**kwargs)
        self.assertEqual(len(initial_result), 1)
        kwargs['original_parameters'] = {'parameter1': 'value1', 'parameter2': 'value2', 'parameter3': 'value3'}
        new_result = results_to_dataframe(**kwargs)
        self.assertEqual(len(new_result), 3)

    def test_results_to_dataframe_returns_dataframe_with_ci_data_when_show_ci_is_true(self) -> None:
        kwargs = self.get_values_for_results_dataframe()
        kwargs['show_ci'] = True
        result = results_to_dataframe(**kwargs)
        self.assertTrue(any('upper' in column_name for column_name in result.columns))

    def test_results_to_dataframe_returns_dataframe_without_ci_data_when_show_ci_is_false(self) -> None:
        kwargs = self.get_values_for_results_dataframe()
        kwargs['show_ci'] = False
        result = results_to_dataframe(**kwargs)
        self.assertTrue(any('upper' not in column_name for column_name in result.columns))


if __name__ == '__main__':
    unittest.main()
